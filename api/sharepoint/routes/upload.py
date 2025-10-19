from __future__ import annotations

import os, io, json, asyncio, inspect, re
import logging
from typing import Optional
from datetime import datetime
from urllib.parse import quote
import httpx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Request

from ..setup import sp_session_logger
from ..schemas import UploadedFile, UploadResponse
from ..graph_auth import get_access_token  # async in your setup

router = APIRouter()
server_logger = logging.getLogger(os.getenv("APP_LOGGER"))

# ------------------------ config ------------------------
GRAPH_BASE = os.getenv("GRAPH_BASE", "https://graph.microsoft.com/v1.0")
GRAPH_DRIVE_ID = os.getenv("GRAPH_DRIVE_ID", "")
GRAPH_ROOT_PREFIX = os.getenv("GRAPH_ROOT_PATH", "")
_INVALID = re.compile(r'[<>:"/\\|?*]')  # Graph/OneDrive forbidden chars


# ------------------------ auth + http helpers ------------------------
async def _token() -> str:
    tok = get_access_token()
    if inspect.isawaitable(tok):
        tok = await tok
    return str(tok)


def _enc(seg: str) -> str:
    return quote(seg.strip("/"), safe="")


def _join(*segs: str) -> str:
    return "/".join(_enc(s) for s in segs if s and s.strip("/"))


async def _graph(
    client: httpx.AsyncClient,
    method: str,
    endpoint: str,
    *,
    json_body=None,
    data=None,
    headers=None,
    timeout=10.0,
    retries=4,
) -> httpx.Response:
    url = GRAPH_BASE + endpoint
    h = {"Authorization": f"Bearer {await _token()}", "Accept": "application/json"}
    if headers:
        h.update(headers)
    for attempt in range(1, retries + 1):
        try:
            sp_session_logger.info(f"[GRAPH] {method} {url} (try {attempt}/{retries})")
            resp = await client.request(
                method, url, headers=h, json=json_body, data=data, timeout=timeout
            )
            sp_session_logger.info(f"[GRAPH] -> {resp.status_code}")
            if resp.status_code in (200, 201, 202, 204):
                return resp
            if resp.status_code in (429, 500, 502, 503, 504) and attempt < retries:
                ra = resp.headers.get("Retry-After")
                delay = float(ra) if ra else 0.8 * (2 ** (attempt - 1))
                logger.warn(f"[GRAPH] retrying in {delay:.2f}s")
                await asyncio.sleep(delay)
                continue
            return resp
        except Exception as e:
            sp_session_logger.error(f"[GRAPH] exception: {e!r}")
            if attempt < retries:
                delay = 0.8 * (2 ** (attempt - 1))
                sp_session_logger.error(f"[GRAPH] retrying in {delay:.2f}s")
                await asyncio.sleep(delay)
                continue
            raise


def sp_safe(name: str) -> str:
    s = _INVALID.sub("-", name)
    return s.strip().rstrip(". ")[:200]


# ------------------------ folder ensure (client + order only) ------------------------
async def _get_by_path(
    client: httpx.AsyncClient, drive_id: str, path: str
) -> httpx.Response:
    ep = f"/drives/{drive_id}/root:/{_join(path)}"
    sp_session_logger.info(f"[FOLDER] GET {ep}")
    return await _graph(client, "GET", ep)


async def ensure_customer_order(
    client: httpx.AsyncClient, drive_id: str, customer: str, order_no: str
) -> tuple[str, bool, bool]:
    root = GRAPH_ROOT_PREFIX.strip("/")
    cust = customer.strip()
    order_name = f"{order_no}.{cust}"
    full_path = "/".join([root, cust, order_name])
    cust_path = "/".join([root, cust])

    r = await _get_by_path(client, drive_id, full_path)
    if r.status_code == 200:
        fid = r.json()["id"]
        sp_session_logger.info(f"[ENSURE] full path exists id={fid}")
        return fid, False, False
    if r.status_code not in (404, 400):
        raise HTTPException(502, f"Check full path failed: {r.text}")

    created_customer = False
    created_order = False

    r2 = await _get_by_path(client, drive_id, cust_path)
    if r2.status_code == 404:
        ep = f"/drives/{drive_id}/root:/{_join(root)}:/children"
        body = {
            "name": cust,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "rename",
        }
        sp_session_logger.info(f"[ENSURE] POST create customer -> {ep}")
        cr = await _graph(client, "POST", ep, json_body=body)
        if cr.status_code not in (200, 201):
            raise HTTPException(502, f"Create customer failed: {cr.text}")
        created_customer = True
    elif r2.status_code != 200:
        raise HTTPException(502, f"Check customer failed: {r2.text}")

    r3 = await _get_by_path(client, drive_id, full_path)
    if r3.status_code == 200:
        fid = r3.json()["id"]
        sp_session_logger.info(f"[ENSURE] order exists id={fid}")
        return fid, created_customer, False

    ep = f"/drives/{drive_id}/root:/{_join(root, cust)}:/children"
    body = {
        "name": order_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "rename",
    }
    sp_session_logger.info(f"[ENSURE] POST create order -> {ep}")
    cr2 = await _graph(client, "POST", ep, json_body=body)
    if cr2.status_code in (200, 201):
        fid = cr2.json()["id"]
        sp_session_logger.info(f"[ENSURE] order created id={fid}")
        return fid, created_customer, True

    r4 = await _get_by_path(client, drive_id, full_path)
    if r4.status_code == 200:
        fid = r4.json()["id"]
        sp_session_logger.info(f"[ENSURE] order found after fallback id={fid}")
        return fid, created_customer, created_order

    raise HTTPException(502, f"Create order failed: {cr2.text}")


# ------------------------ XLSX builder (minimal checklist -> workbook) ------------------------
def build_qc_xlsx_from_checklist(order_no: str, checklist: dict) -> io.BytesIO:
    """
    checklist schema (minimal):
    {
      "items": [
        { "code": "13502.06.01",
          "checks": { "Structure_Weld": "pass", ... },
          "comments": { "Structure_Weld": "note...", ... }
        },
        ...
      ]
    }
    """
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    items = checklist.get("items", []) if isinstance(checklist, dict) else []
    if not items:
        ws = wb.create_sheet("Summary")
        ws["A1"] = f"No checklist items for order {order_no}"
    else:
        for idx, raw_item in enumerate(items, start=1):
            item = raw_item or {}
            code = str(item.get("code") or "").strip() or "no-code"
            ws = wb.create_sheet((f"Item {idx} - {code}")[:31])

            checks = item.get("checks") or {}
            comments = item.get("comments") or {}

            # One table: Inspection Item | Judge | Comment
            ws.append(["Inspection Item", "Judge", "Comments"])

            # Preserve check order; then any comment-only rows
            keys = list(checks.keys()) + [k for k in comments.keys() if k not in checks]

            for k in keys:
                ws.append([str(k), str(checks.get(k, "")), str(comments.get(k, ""))])

            # autosize
            for col in range(1, ws.max_column + 1):
                letter = get_column_letter(col)
                width = (
                    max(
                        (len(str(c.value)) if c.value is not None else 0)
                        for c in ws[letter]
                    )
                    + 2
                )
                ws.column_dimensions[letter].width = min(max(12, width), 60)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ------------------------ uploads ------------------------
def _mime_from_upload(up: UploadFile) -> str:
    if up.content_type:
        return up.content_type
    name = (up.filename or "").lower()
    if name.endswith(".jpg") or name.endswith(".jpeg"):
        return "image/jpeg"
    if name.endswith(".png"):
        return "image/png"
    return "application/octet-stream"


async def put_small_file(
    client: httpx.AsyncClient,
    drive_id: str,
    dest_path_with_name: str,
    data: bytes,
    mime: str,
) -> dict:
    ep = f"/drives/{drive_id}/root:/{_join(dest_path_with_name)}:/content?@microsoft.graph.conflictBehavior=rename"
    sp_session_logger.info(f"[UPLOAD] small PUT {ep} bytes={len(data)} mime={mime}")
    r = await _graph(
        client, "PUT", ep, data=data, headers={"Content-Type": mime}, timeout=30.0
    )
    if r.status_code not in (200, 201):
        raise HTTPException(502, f"PUT small failed: {r.text}")
    return r.json()


async def create_upload_session(
    client: httpx.AsyncClient, drive_id: str, dest_path_with_name: str
) -> str:
    ep = f"/drives/{drive_id}/root:/{_join(dest_path_with_name)}:/createUploadSession"
    body = {"@microsoft.graph.conflictBehavior": "rename", "deferCommit": False}
    r = await _graph(client, "POST", ep, json_body=body)
    if r.status_code not in (200, 201):
        raise HTTPException(502, f"createUploadSession failed: {r.text}")
    return r.json()["uploadUrl"]


async def upload_large_file(
    client: httpx.AsyncClient,
    drive_id: str,
    dest_path_with_name: str,
    up: UploadFile,
    mime: str,
) -> dict:
    # Simple size detection and chunk loop using in-memory bytes for clarity
    await up.seek(0)
    data = await up.read()
    size = len(data)
    await up.seek(0)
    upload_url = await create_upload_session(client, drive_id, dest_path_with_name)
    chunk = 8 * 1024 * 1024
    sent = 0
    part = 0
    sp_session_logger.info(
        f"[UPLOAD] large begin size={size} chunk={chunk} mime={mime}"
    )
    while sent < size:
        end = min(sent + chunk, size)
        piece = data[sent:end]
        part += 1
        headers = {
            "Content-Length": str(len(piece)),
            "Content-Range": f"bytes {sent}-{end-1}/{size}",
            "Content-Type": mime,
        }
        sp_session_logger.info(f"[UPLOAD]   part {part} {sent}-{end-1}/{size}")
        resp = await client.put(
            upload_url, headers=headers, content=piece, timeout=60.0
        )
        if resp.status_code in (200, 201):
            sp_session_logger.info("[UPLOAD] large complete")
            return resp.json()
        if resp.status_code == 202:
            sent = end
            continue
        raise HTTPException(
            502, f"Resumable chunk failed: {resp.status_code} {resp.text}"
        )
    raise HTTPException(502, "Resumable upload ended unexpectedly")


# ---------------------------------- route ----------------------------------- #
@router.post("/upload", response_model=UploadResponse)
async def upload_qc(
    request: Request,
    orderNo: str = Form(...),
    customerName: str = Form(..., alias="client"),
    checklist: Optional[str] = Form(None),
    folderId: Optional[str] = Form(None),
    files: list[UploadFile] = File(...),
    fileSignal: Optional[str] = Form(None),
):
    ct = request.headers.get("content-type", "")
    if not ct.startswith("multipart/form-data"):
        raise HTTPException(
            status_code=415,
            detail="Send as multipart/form-data; do not set Content-Type manually.",
        )

    if not files:  # allow 'files[]' too
        form = await request.form()
        files = form.getlist("files[]")  # returns UploadFile objects

    customerName = sp_safe(customerName)
    if fileSignal == "first":
        sp_session_logger.info(f"=== QC SESSION: {orderNo}.{customerName} ===")
        sp_session_logger.info(
            f"[REQ] orderNo='{orderNo}' client='{customerName}' files={len(files)} checklist_len={len(checklist or '')}"
        )

    async with httpx.AsyncClient() as client:
        if folderId:
            created_customer = created_order = False
        else:
            try:
                folderId, created_customer, created_order = await ensure_customer_order(
                    client, GRAPH_DRIVE_ID, customerName, orderNo
                )
                sp_session_logger.info(
                    f"[ENSURE] id={folderId} created_customer={created_customer} created_order={created_order}"
                )
            except HTTPException as e:
                sp_session_logger.error(f"[ERR] ensure failed: {e.detail}")
                raise HTTPException(502, f"Folder ensure failed: {e.detail}")

        # upload XLSX once if checklist provided
        if checklist and checklist.strip() not in ("", "null", "{}"):
            try:
                chk = json.loads(checklist)
                xlsx = build_qc_xlsx_from_checklist(orderNo, chk)
                xname = f"{orderNo}_QC_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
                dest = "/".join(
                    [
                        GRAPH_ROOT_PREFIX.strip("/"),
                        customerName.strip(),
                        f"{orderNo}.{customerName.strip()}",
                        xname,
                    ]
                )
                data = xlsx.getvalue()
                # most manifests are small; use small PUT
                _ = await put_small_file(
                    client,
                    GRAPH_DRIVE_ID,
                    dest,
                    data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                sp_session_logger.info(f"[XLSX] uploaded '{xname}'")
            except Exception as e:
                sp_session_logger.error(f"[XLSX] skipped due to error: {e!s}")

        # upload photos one-by-one (client can batch a few for (i/N) progress)
        uploaded: list[UploadedFile] = []
        for i, up in enumerate(files, start=1):
            base = up.filename or f"photo_{i}.jpg"
            try:
                await up.seek(0)
                buf = await up.read()
                size = len(buf)
                await up.seek(0)
                mime = _mime_from_upload(up)
                dest_rel = "/".join(
                    [
                        GRAPH_ROOT_PREFIX.strip("/"),
                        customerName.strip(),
                        f"{orderNo}.{customerName.strip()}",
                        base,
                    ]
                )
                if size <= 4 * 1024 * 1024:
                    meta = await put_small_file(
                        client, GRAPH_DRIVE_ID, dest_rel, buf, mime
                    )
                else:
                    meta = await upload_large_file(
                        client, GRAPH_DRIVE_ID, dest_rel, up, mime
                    )
                uploaded.append(
                    UploadedFile(
                        id=meta.get("id"),
                        name=meta.get("name", base),
                        webUrl=meta.get("webUrl"),
                        size=meta.get("size", size),
                        content_type=mime,
                    )
                )
                sp_session_logger.info(f"[OK] {base} -> {meta.get('webUrl')}")
            except Exception as e:
                sp_session_logger.error(f"[ERR] file '{base}' failed: {e!s}")
                uploaded.append(
                    UploadedFile(name=base, webUrl=None, size=0, content_type=None)
                )

        if fileSignal == "eof":
            sp_session_logger.info("")  # signal for log spacing

        ok_count = sum(1 for u in uploaded if u.id)
        return UploadResponse(
            ok=ok_count == len(uploaded) and len(uploaded) > 0,
            customer=customerName,
            order_no=orderNo,
            folderId=folderId,
            created_customer=created_customer,
            created_order=created_order,
            uploaded_count=ok_count,
            uploaded=uploaded,
        )
